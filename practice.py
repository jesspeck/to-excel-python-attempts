'''
pathtest = path
filename = "testoutput.xlsx"
path = os.path.join(path, y + 'vs' + z + '_output.xlsx')
#df.to_csv(os.path.join(path, 'out.csv'), index=True) 
# TODO: clean up unused code :P
# export dataframes to excel sheets
import openpyxl
# from openpyxl import Workbook
# from openpyxl.cell.cell import WriteOnlyCell
# from openpyxl.utils.dataframe import dataframe_to_rows
# wb = Workbook()
# ws = wb.active

result_df1 = result_df.set_index('query')

import pandas.io.formats.excel

df_test = pd.DataFrame([
    "simple string",
    "20"
])

from openpyxl import Workbook, load_workbook

#%%
df_test.to_excel(os.path.join(pathtest, filename))

# df_test.to_excel(pathtest2, header=False, index=False)


# pd.io.formats.style.Styler.format.header_style = None

with pd.ExcelWriter(path, engine='xlsxwriter', engine_kwargs={'options': {'strings_to_urls': False, 'strings_to_formulas': False}}) as writer:
    # writer.book = book
    # writer = pd.ExcelWriter(output, engine='openpyxl')
    result_df1.to_excel(writer, sheet_name="Sheet1")
    df1.to_excel(writer, sheet_name="Sheet2")
    df2.to_excel(writer, sheet_name="Sheet3")
    

# writer.save()
# book.close()
# writer.close()
# writer.handles = None

"""
wb = Workbook(write_only=True)
ws1 = wb.create_sheet()
ws1.title = "results"
cell = WriteOnlyCell(ws1)

def format_first_row(row, cell):
    for c in row:
        cell.value = c
        yield cell

rows = dataframe_to_rows(result_df)
first_row = format_first_row(next(rows), cell)
ws1.append(first_row)

for row in rows:
    row = list(row)
    cell.value = row[0]
    row[0] = cell
    ws1.append(row)

# wb.save("openpyxl_stream.xlsx")

ws2 = wb.create_sheet(title=y)

for r in dataframe_to_rows(df1, index=True, header=True):
    ws2.append(r)
ws3 = wb.create_sheet(title=z)
for r in dataframe_to_rows(df2, index=True, header=True):
    ws3.append(r)
writer = pd.ExcelWriter(path, engine='xlsxwriter')

result_df1.to_excel(writer, sheet_name='results')
df1.to_excel(writer, sheet_name='inputdate1')
df2.to_excel(writer, sheet_name='inputdate2')
"""
# writer.save(path)
# wb.save(filename = path)

# DONE fix excel file output
# for r in dataframe_to_rows(result_df, index=True, header=True):
#     ws.append(r)

#ws1 = wb.create_sheet("Mysheet") # create a sheet
'''
